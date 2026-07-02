"""
Modulo Economico STEELEX — struttura semplificata:
  - Computo: voci di costo + ricarico → preventivo cliente
  - Spese:   registro semplice con foto/PDF allegato
  - SAL:     fatturazione cliente a milestone
  - Gantt:   fasi di lavoro con cronoprogramma
"""
import io
import os
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional, Any
from datetime import date, datetime
from app.database import get_db
from app.models.economico import (
    PreventivoCantiere, SAL, StatoSAL, StatoPreventivo,
    Spesa, CategoriaSpesa, FaseLavoro, StatoFase,
    OrdineAcquisto, FatturaFornitore, BollaConsegna  # mantenuti per compatibilità dati esistenti
)
from app.models.cantiere import Cantiere
from app.models.utente import RuoloUtente, Utente
from app.auth import get_current_user
from app.storage import salva_file
from app.routers.notifiche import notifica_cantiere
from pydantic import BaseModel

router = APIRouter(prefix="/cantieri", tags=["Economico"])


# ─── AUTORIZZAZIONI ───────────────────────────────────────────────────────────

_RUOLI_STAFF = (
    RuoloUtente.admin,
    RuoloUtente.capo_cantiere,
    RuoloUtente.capo_cantiere_sub,
    RuoloUtente.direzione_lavori,
    RuoloUtente.amministrazione,
)

def _check(cantiere_id: int, db: Session, user: Utente) -> Cantiere:
    """Verifica accesso al cantiere (NON implica accesso all'economia)."""
    c = db.query(Cantiere).filter(Cantiere.id == cantiere_id).first()
    if not c:
        raise HTTPException(404, "Cantiere non trovato")
    if user.ruolo == RuoloUtente.admin:
        return c
    if user.ruolo == RuoloUtente.amministrazione:
        return c  # amministrazione vede tutti i cantieri
    if user.ruolo == RuoloUtente.capo_cantiere and c.responsabile_id == user.id:
        return c
    # capo_cantiere_sub, direzione_lavori, artigiano, fornitore, cliente:
    # accesso solo se assegnati al cantiere
    if user.id in [u.id for u in c.artigiani]:
        return c
    raise HTTPException(403, "Accesso negato al cantiere")

def _solo_staff(user: Utente):
    """Operazioni di scrittura Gantt/SAL/fasi — tutti i ruoli cantiere eccetto cliente."""
    if user.ruolo not in _RUOLI_STAFF:
        raise HTTPException(403, "Accesso riservato allo staff di cantiere")

def _solo_economia(user: Utente):
    """Modulo economico (preventivi, spese, SAL) — admin, capo cantiere e amministrazione."""
    if user.ruolo not in (RuoloUtente.admin, RuoloUtente.capo_cantiere, RuoloUtente.amministrazione):
        raise HTTPException(403, "Modulo economico riservato ad admin, capo cantiere e amministrazione")

def _economia_o_dl(user: Utente):
    """Lettura economia — admin, capo_cantiere, amministrazione + direzione_lavori."""
    if user.ruolo not in (RuoloUtente.admin, RuoloUtente.capo_cantiere, RuoloUtente.amministrazione, RuoloUtente.direzione_lavori):
        raise HTTPException(403, "Accesso non autorizzato")

def _is_dl(user: Utente) -> bool:
    return user.ruolo == RuoloUtente.direzione_lavori

def _filtra_voce_per_dl(voce: dict) -> dict:
    """Rimuove dati di costo interno — il DL vede solo prezzi cliente."""
    return {k: v for k, v in voce.items() if k not in ('costo_unitario', 'ricarico_perc', 'totale_costo')}

def _preventivo_per_dl(prev) -> dict:
    """Restituisce il preventivo con voci filtrate per il DL."""
    d = {c.name: getattr(prev, c.name) for c in prev.__table__.columns}
    d['voci'] = [_filtra_voce_per_dl(v) for v in (prev.voci or [])]
    d['costo_totale'] = None
    return d

def _blocca_cliente(user: Utente):
    if user.ruolo == RuoloUtente.cliente:
        raise HTTPException(403, "Dati economici non accessibili al cliente")


# ─── RIEPILOGO ────────────────────────────────────────────────────────────────

class RiepilogoOut(BaseModel):
    budget_preventivo: float     # totale preventivo accettato (IVA esclusa)
    budget_iva: float            # con IVA
    totale_speso: float          # somma spese registrate
    margine_atteso: float        # budget - totale_speso
    totale_sal_emessi: float     # SAL emessi + pagati
    totale_sal_pagati: float     # SAL incassati
    da_incassare: float
    spese_per_categoria: dict

@router.get("/{cantiere_id}/economia")
def riepilogo(cantiere_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user)
    _economia_o_dl(user)

    preventivi = db.query(PreventivoCantiere).filter(PreventivoCantiere.cantiere_id == cantiere_id).all()
    spese = db.query(Spesa).filter(Spesa.cantiere_id == cantiere_id).all()
    sal_list = db.query(SAL).filter(SAL.cantiere_id == cantiere_id).all()
    cantiere = db.query(Cantiere).filter(Cantiere.id == cantiere_id).first()

    # Usa i preventivi accettati; se nessuno accettato, somma tutti
    prev_accettati = [p for p in preventivi if p.stato == "accettato"]
    prev_base = prev_accettati if prev_accettati else preventivi
    budget     = sum(p.subtotale for p in prev_base) or (cantiere.budget or 0)
    budget_iva = sum(p.totale    for p in prev_base) or budget

    totale_speso = sum(s.importo for s in spese)
    sal_emessi = sum(s.importo for s in sal_list if s.stato in ("emesso","pagato"))
    sal_pagati = sum(s.importo for s in sal_list if s.stato == "pagato")

    cat_totali = {}
    for s in spese:
        cat = s.categoria or "altro"
        cat_totali[cat] = cat_totali.get(cat, 0) + s.importo

    if _is_dl(user):
        return RiepilogoOut(
            budget_preventivo=budget,
            budget_iva=budget_iva,
            totale_speso=0,
            margine_atteso=0,
            totale_sal_emessi=sal_emessi,
            totale_sal_pagati=sal_pagati,
            da_incassare=sal_emessi - sal_pagati,
            spese_per_categoria={},
        )
    return RiepilogoOut(
        budget_preventivo=budget,
        budget_iva=budget_iva,
        totale_speso=totale_speso,
        margine_atteso=budget - totale_speso,
        totale_sal_emessi=sal_emessi,
        totale_sal_pagati=sal_pagati,
        da_incassare=sal_emessi - sal_pagati,
        spese_per_categoria=cat_totali,
    )


# ─── COMPUTO / PREVENTIVO ─────────────────────────────────────────────────────

class PreventivoOut(BaseModel):
    id: int; cantiere_id: int; numero: Optional[str]; data: Optional[date]
    validita_giorni: int; voci: Any; subtotale: float; costo_totale: float
    iva_perc: float; totale: float; acconto_perc: float; acconto_importo: float
    acconto_ricevuto: float; data_acconto: Optional[date]; stato: str
    pdf_url: Optional[str]; note: Optional[str]; creato_il: Optional[datetime]
    class Config: from_attributes = True

class PreventivoCreate(BaseModel):
    numero: Optional[str] = None
    data_preventivo: Optional[date] = None
    validita_giorni: int = 30
    voci: list = []
    iva_perc: float = 22.0
    acconto_perc: float = 30.0
    note: Optional[str] = None

class PreventivoUpdate(BaseModel):
    numero: Optional[str] = None
    data_preventivo: Optional[date] = None
    voci: Optional[list] = None
    iva_perc: Optional[float] = None
    acconto_perc: Optional[float] = None
    acconto_ricevuto: Optional[float] = None
    data_acconto: Optional[date] = None
    stato: Optional[str] = None
    note: Optional[str] = None

def _ricalcola(prev, voci, iva_perc, acconto_perc):
    subtotale = sum(v.get("totale_cliente", 0) for v in voci)
    costo_totale = sum(v.get("totale_costo", 0) for v in voci)
    totale = round(subtotale * (1 + iva_perc / 100), 2)
    prev.voci = voci
    prev.subtotale = round(subtotale, 2)
    prev.costo_totale = round(costo_totale, 2)
    prev.iva_perc = iva_perc
    prev.totale = totale
    prev.acconto_perc = acconto_perc
    prev.acconto_importo = round(totale * acconto_perc / 100, 2)

@router.get("/{cantiere_id}/preventivi")
def lista_preventivi(cantiere_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _economia_o_dl(user)
    records = db.query(PreventivoCantiere).filter(PreventivoCantiere.cantiere_id == cantiere_id).order_by(PreventivoCantiere.creato_il.desc()).all()
    changed = False
    for prev in records:
        voci = prev.voci or []
        if prev.totale == 0 and len(voci) > 0:
            _ricalcola(prev, voci, prev.iva_perc, prev.acconto_perc)
            changed = True
    if changed:
        try: db.commit()
        except Exception: db.rollback()
    if _is_dl(user):
        return [_preventivo_per_dl(p) for p in records]
    return records

@router.post("/{cantiere_id}/preventivi", response_model=PreventivoOut, status_code=201)
def crea_preventivo(cantiere_id: int, body: PreventivoCreate, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    try:
        prev = PreventivoCantiere(
            cantiere_id=cantiere_id,
            numero=body.numero,
            data=body.data_preventivo,
            validita_giorni=body.validita_giorni,
            iva_perc=body.iva_perc,
            acconto_perc=body.acconto_perc,
            note=body.note,
        )
        _ricalcola(prev, body.voci, body.iva_perc, body.acconto_perc)
        db.add(prev); db.commit(); db.refresh(prev)
        try:
            notifica_cantiere(db, cantiere_id,
                ruoli=["admin", "direzione_lavori"],
                titolo="💰 Nuovo preventivo creato",
                corpo=f"{user.nome} {user.cognome}: preventivo n°{body.numero}",
                escludi_id=user.id,
                url=f"/cantieri/{cantiere_id}#economia",
            )
        except Exception: pass
        return prev
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Errore: {e}")

@router.put("/{cantiere_id}/preventivi/{prev_id}", response_model=PreventivoOut)
def aggiorna_preventivo(cantiere_id: int, prev_id: int, body: PreventivoUpdate, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    prev = db.query(PreventivoCantiere).filter(PreventivoCantiere.id == prev_id, PreventivoCantiere.cantiere_id == cantiere_id).first()
    if not prev: raise HTTPException(404, "Non trovato")
    upd = body.model_dump(exclude_none=True)
    if "data_preventivo" in upd: prev.data = upd.pop("data_preventivo")
    voci = upd.pop("voci", prev.voci or [])
    iva = upd.pop("iva_perc", prev.iva_perc)
    acc = upd.pop("acconto_perc", prev.acconto_perc)
    for k, v in upd.items(): setattr(prev, k, v)
    _ricalcola(prev, voci, iva, acc)
    db.commit(); db.refresh(prev)
    return prev

class VoceExtraBody(BaseModel):
    descrizione: str
    importo: float = 0.0
    categoria: str = "Altro"
    um: str = "corpo"
    ricarico_perc: float = 0.0

@router.post("/{cantiere_id}/preventivi/{prev_id}/voce-extra", response_model=PreventivoOut)
def aggiungi_voce_extra(
    cantiere_id: int, prev_id: int, body: VoceExtraBody,
    db: Session = Depends(get_db), user: Utente = Depends(get_current_user),
):
    """Aggiunge una voce extra preventivo al computo senza riaprire il form completo."""
    from sqlalchemy.orm.attributes import flag_modified
    _check(cantiere_id, db, user); _solo_economia(user)
    prev = db.query(PreventivoCantiere).filter(
        PreventivoCantiere.id == prev_id, PreventivoCantiere.cantiere_id == cantiere_id
    ).first()
    if not prev:
        raise HTTPException(404, "Preventivo non trovato")
    costo = round(body.importo, 2)
    prezzo = round(costo * (1 + body.ricarico_perc / 100), 2)
    voce = {
        "id": int(datetime.now().timestamp() * 1000),
        "descrizione": f"⚠ EXTRA — {body.descrizione}",
        "categoria": body.categoria,
        "qt": 1,
        "um": body.um,
        "costo_unitario": costo,
        "ricarico_perc": round(body.ricarico_perc, 2),
        "prezzo_unitario": prezzo,
        "totale_costo": costo,
        "totale_cliente": prezzo,
    }
    voci = list(prev.voci or [])
    voci.append(voce)
    _ricalcola(prev, voci, prev.iva_perc, prev.acconto_perc)
    flag_modified(prev, "voci")
    db.commit(); db.refresh(prev)
    try:
        importo_str = f"€ {prezzo:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        notifica_cantiere(db, cantiere_id,
            ruoli=["direzione_lavori"],
            titolo="⚠️ Extra preventivo aggiunto al computo",
            corpo=f"{body.descrizione[:80]} — {importo_str} (prezzo cliente)",
            escludi_id=user.id,
            tipo="extra_preventivo",
            url=f"/cantieri/{cantiere_id}#economia",
        )
    except Exception: pass
    if _is_dl(user):
        return _preventivo_per_dl(prev)
    return prev

@router.post("/{cantiere_id}/preventivi/{prev_id}/pdf", response_model=PreventivoOut)
async def upload_pdf_preventivo(cantiere_id: int, prev_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    prev = db.query(PreventivoCantiere).filter(PreventivoCantiere.id == prev_id, PreventivoCantiere.cantiere_id == cantiere_id).first()
    if not prev: raise HTTPException(404, "Non trovato")
    ext = os.path.splitext(file.filename or ".pdf")[1].lower() or ".pdf"
    url, _ = salva_file(await file.read(), f"preventivi/{cantiere_id}", ext)
    prev.pdf_url = url; db.commit(); db.refresh(prev)
    return prev

@router.delete("/{cantiere_id}/preventivi/{prev_id}", status_code=204)
def elimina_preventivo(cantiere_id: int, prev_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    prev = db.query(PreventivoCantiere).filter(PreventivoCantiere.id == prev_id, PreventivoCantiere.cantiere_id == cantiere_id).first()
    if not prev: raise HTTPException(404, "Non trovato")
    db.delete(prev); db.commit()


# ─── SPESE ────────────────────────────────────────────────────────────────────

class SpesaOut(BaseModel):
    id: int; cantiere_id: int; descrizione: str; fornitore: Optional[str]
    categoria: str; importo: float; data: Optional[date]; note: Optional[str]
    allegato_url: Optional[str]; allegato_tipo: Optional[str]; creato_il: Optional[datetime]
    class Config: from_attributes = True

class SpesaCreate(BaseModel):
    descrizione: str
    fornitore: Optional[str] = None
    categoria: str = "materiali"
    importo: float
    data: Optional[date] = None
    note: Optional[str] = None

class SpesaUpdate(BaseModel):
    descrizione: Optional[str] = None
    fornitore: Optional[str] = None
    categoria: Optional[str] = None
    importo: Optional[float] = None
    data: Optional[date] = None
    note: Optional[str] = None

@router.get("/{cantiere_id}/spese", response_model=List[SpesaOut])
def lista_spese(cantiere_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    return db.query(Spesa).filter(Spesa.cantiere_id == cantiere_id).order_by(Spesa.creato_il.desc()).all()

@router.post("/{cantiere_id}/spese", response_model=SpesaOut, status_code=201)
def registra_spesa(cantiere_id: int, body: SpesaCreate, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    try:
        s = Spesa(cantiere_id=cantiere_id, creato_da=user.id, **body.model_dump())
        db.add(s); db.commit(); db.refresh(s)
        try:
            notifica_cantiere(db, cantiere_id,
                ruoli=["admin", "direzione_lavori"],
                titolo="🧾 Nuova spesa registrata",
                corpo=f"{user.nome} {user.cognome}: {getattr(body,'descrizione','')[:60]} — €{getattr(body,'importo','')}",
                escludi_id=user.id,
                url=f"/cantieri/{cantiere_id}#economia",
            )
        except Exception: pass
        return s
    except Exception as e:
        db.rollback(); raise HTTPException(500, f"Errore: {e}")

@router.put("/{cantiere_id}/spese/{spesa_id}", response_model=SpesaOut)
def aggiorna_spesa(cantiere_id: int, spesa_id: int, body: SpesaUpdate, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    s = db.query(Spesa).filter(Spesa.id == spesa_id, Spesa.cantiere_id == cantiere_id).first()
    if not s: raise HTTPException(404, "Non trovata")
    for k, v in body.model_dump(exclude_none=True).items(): setattr(s, k, v)
    db.commit(); db.refresh(s)
    return s

@router.post("/{cantiere_id}/spese/{spesa_id}/allegato", response_model=SpesaOut)
async def upload_allegato_spesa(cantiere_id: int, spesa_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    s = db.query(Spesa).filter(Spesa.id == spesa_id, Spesa.cantiere_id == cantiere_id).first()
    if not s: raise HTTPException(404, "Non trovata")
    ext = os.path.splitext(file.filename or "")[1].lower()
    contenuto = await file.read()
    url, _ = salva_file(contenuto, f"spese/{cantiere_id}", ext)
    s.allegato_url = url
    s.allegato_tipo = "pdf" if ext == ".pdf" else "foto"
    db.commit(); db.refresh(s)
    return s

@router.delete("/{cantiere_id}/spese/{spesa_id}", status_code=204)
def elimina_spesa(cantiere_id: int, spesa_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    s = db.query(Spesa).filter(Spesa.id == spesa_id, Spesa.cantiere_id == cantiere_id).first()
    if not s: raise HTTPException(404, "Non trovata")
    db.delete(s); db.commit()


# ─── AUTORIZZAZIONE FATTURE ───────────────────────────────────────────────────

@router.get("/{cantiere_id}/fatture", response_model=List[dict])
def lista_fatture(cantiere_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    fatture = db.query(FatturaFornitore).filter(FatturaFornitore.cantiere_id == cantiere_id)\
                .order_by(FatturaFornitore.data_fattura.desc()).all()
    result = []
    for f in fatture:
        aut = None
        if getattr(f, 'autorizzata_da', None):
            u = db.query(Utente).filter(Utente.id == f.autorizzata_da).first()
            aut = f"{u.nome} {u.cognome}".strip() if u else None
        result.append({
            "id": f.id, "fornitore_nome": f.fornitore_nome, "numero_fattura": f.numero_fattura,
            "descrizione": f.descrizione, "importo_netto": f.importo_netto,
            "importo_totale": f.importo_totale, "data_fattura": str(f.data_fattura) if f.data_fattura else None,
            "data_scadenza": str(f.data_scadenza) if f.data_scadenza else None,
            "stato": f.stato.value if f.stato else None, "pdf_url": f.pdf_url,
            "autorizzata": getattr(f, 'autorizzata', False) or False,
            "autorizzata_da_nome": aut,
            "autorizzata_il": str(getattr(f, 'autorizzata_il', None)) if getattr(f, 'autorizzata_il', None) else None,
        })
    return result


@router.post("/{cantiere_id}/fatture", response_model=dict, status_code=201)
def crea_fattura(cantiere_id: int, body: dict, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    from app.models.economico import StatoFattura
    try: stato = StatoFattura(body.get("stato", "ricevuta"))
    except ValueError: stato = StatoFattura.ricevuta
    f = FatturaFornitore(
        cantiere_id=cantiere_id, fornitore_nome=body["fornitore_nome"],
        numero_fattura=body.get("numero_fattura"), descrizione=body.get("descrizione"),
        importo_netto=body.get("importo_netto", 0), iva_perc=body.get("iva_perc", 22),
        importo_iva=body.get("importo_iva", 0), importo_totale=body.get("importo_totale", 0),
        data_fattura=body.get("data_fattura") or None,
        data_scadenza=body.get("data_scadenza") or None, stato=stato,
    )
    db.add(f); db.commit(); db.refresh(f)
    return {"id": f.id, "messaggio": "Fattura creata"}


@router.post("/{cantiere_id}/fatture/{fattura_id}/autorizza", response_model=dict)
def autorizza_fattura(cantiere_id: int, fattura_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    from datetime import datetime as dt
    f = db.query(FatturaFornitore).filter(FatturaFornitore.id == fattura_id, FatturaFornitore.cantiere_id == cantiere_id).first()
    if not f: raise HTTPException(404, "Fattura non trovata")
    from sqlalchemy import text as _t
    db.execute(_t("UPDATE fatture_fornitori SET autorizzata=TRUE, autorizzata_da=:uid, autorizzata_il=NOW() WHERE id=:fid"),
               {"uid": user.id, "fid": fattura_id})
    db.commit()
    return {"messaggio": "Fattura autorizzata", "autorizzata_da": f"{user.nome} {user.cognome}".strip()}


@router.delete("/{cantiere_id}/fatture/{fattura_id}", status_code=204)
def elimina_fattura(cantiere_id: int, fattura_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    f = db.query(FatturaFornitore).filter(FatturaFornitore.id == fattura_id, FatturaFornitore.cantiere_id == cantiere_id).first()
    if not f: raise HTTPException(404)
    db.delete(f); db.commit()


# ─── SAL ──────────────────────────────────────────────────────────────────────

class SALOut(BaseModel):
    id: int; cantiere_id: int; numero: int; titolo: str
    percentuale: float; importo: float; data: Optional[date]
    stato: str; note: Optional[str]; creato_il: Optional[datetime]
    class Config: from_attributes = True

class SALCreate(BaseModel):
    titolo: str; percentuale: float; importo: float
    data: Optional[date] = None; stato: str = "bozza"; note: Optional[str] = None

class SALUpdate(BaseModel):
    titolo: Optional[str] = None; percentuale: Optional[float] = None
    importo: Optional[float] = None; data: Optional[date] = None
    stato: Optional[str] = None; note: Optional[str] = None

@router.get("/{cantiere_id}/sal", response_model=List[SALOut])
def lista_sal(cantiere_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    return db.query(SAL).filter(SAL.cantiere_id == cantiere_id).order_by(SAL.numero).all()

@router.post("/{cantiere_id}/sal", response_model=SALOut, status_code=201)
def crea_sal(cantiere_id: int, body: SALCreate, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    ultimo = db.query(SAL).filter(SAL.cantiere_id == cantiere_id).order_by(SAL.numero.desc()).first()
    sal = SAL(cantiere_id=cantiere_id, numero=(ultimo.numero + 1 if ultimo else 1), **body.model_dump())
    db.add(sal); db.commit(); db.refresh(sal)
    try:
        notifica_cantiere(db, cantiere_id,
            ruoli=["admin", "direzione_lavori"],
            titolo="📊 Nuovo SAL emesso",
            corpo=f"{user.nome} {user.cognome}: SAL n°{sal.numero} — {getattr(body,'titolo','')[:60]}",
            escludi_id=user.id,
            url=f"/cantieri/{cantiere_id}#economia",
        )
    except Exception: pass
    return sal

@router.put("/{cantiere_id}/sal/{sal_id}", response_model=SALOut)
def aggiorna_sal(cantiere_id: int, sal_id: int, body: SALUpdate, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    sal = db.query(SAL).filter(SAL.id == sal_id, SAL.cantiere_id == cantiere_id).first()
    if not sal: raise HTTPException(404, "Non trovato")
    for k, v in body.model_dump(exclude_none=True).items(): setattr(sal, k, v)
    db.commit(); db.refresh(sal)
    return sal

@router.delete("/{cantiere_id}/sal/{sal_id}", status_code=204)
def elimina_sal(cantiere_id: int, sal_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_economia(user)
    sal = db.query(SAL).filter(SAL.id == sal_id, SAL.cantiere_id == cantiere_id).first()
    if not sal: raise HTTPException(404, "Non trovato")
    db.delete(sal); db.commit()


# ─── GANTT / FASI ─────────────────────────────────────────────────────────────

class FaseOut(BaseModel):
    id: int; cantiere_id: int; sal_id: Optional[int]; nome: str
    categoria: str; colore: str; ordine: int; data_inizio: Optional[date]
    data_fine_prevista: Optional[date]; data_fine_reale: Optional[date]
    percentuale: float; stato: str; note: Optional[str]; creato_il: Optional[datetime]
    visibile_cliente: bool = False
    class Config: from_attributes = True

class FaseCreate(BaseModel):
    nome: str; categoria: str = "lavorazione"; colore: str = "#FF6B00"
    ordine: int = 0; data_inizio: Optional[date] = None
    data_fine_prevista: Optional[date] = None; sal_id: Optional[int] = None
    percentuale: float = 0.0; stato: str = "pianificata"; note: Optional[str] = None
    visibile_cliente: bool = False

class FaseUpdate(BaseModel):
    nome: Optional[str] = None; categoria: Optional[str] = None
    colore: Optional[str] = None; ordine: Optional[int] = None
    data_inizio: Optional[date] = None; data_fine_prevista: Optional[date] = None
    data_fine_reale: Optional[date] = None; percentuale: Optional[float] = None
    stato: Optional[str] = None; sal_id: Optional[int] = None; note: Optional[str] = None
    visibile_cliente: Optional[bool] = None

@router.get("/{cantiere_id}/fasi", response_model=List[FaseOut])
def lista_fasi(cantiere_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user)
    return db.query(FaseLavoro).filter(FaseLavoro.cantiere_id == cantiere_id).order_by(FaseLavoro.ordine, FaseLavoro.data_inizio).all()

@router.get("/{cantiere_id}/aggiornamenti-cliente")
def aggiornamenti_cliente(cantiere_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    """Dati aggregati per la pagina aggiornamenti visibile al cliente."""
    _check(cantiere_id, db, user)
    from app.models.diario import DiarioGiornaliero
    from app.models.cantiere import Cantiere as CantiereModel
    import datetime

    cantiere = db.query(CantiereModel).filter(CantiereModel.id == cantiere_id).first()
    fasi = db.query(FaseLavoro).filter(FaseLavoro.cantiere_id == cantiere_id).order_by(FaseLavoro.ordine, FaseLavoro.data_inizio).all()

    # Fasi visibili al cliente (solo quelle segnate)
    fasi_da_mostrare = [f for f in fasi if f.visibile_cliente]

    # Avanzamento globale calcolato su TUTTE le fasi (non solo quelle condivise)
    perc_globale = round(sum(f.percentuale for f in fasi) / len(fasi), 1) if fasi else 0

    # Note diario condivise con cliente (ultime 8)
    note = (db.query(DiarioGiornaliero)
              .filter(DiarioGiornaliero.cantiere_id == cantiere_id,
                      DiarioGiornaliero.condividi_cliente == True)
              .order_by(DiarioGiornaliero.data.desc())
              .limit(8).all())

    # Prossimi appuntamenti = fasi visibili con data_inizio futura o oggi
    oggi = datetime.date.today()
    appuntamenti = [
        f for f in fasi_da_mostrare
        if f.data_inizio and f.data_inizio >= oggi and f.stato not in ("completata",)
    ]
    appuntamenti.sort(key=lambda f: f.data_inizio)

    return {
        "cantiere_nome": cantiere.nome if cantiere else "",
        "avanzamento_globale": perc_globale,
        "totale_fasi": len(fasi),
        "fasi_condivise": len(fasi_da_mostrare),
        "fasi": [
            {"id": f.id, "nome": f.nome, "categoria": f.categoria, "colore": f.colore,
             "percentuale": f.percentuale, "stato": f.stato,
             "visibile_cliente": f.visibile_cliente,
             "data_inizio": str(f.data_inizio) if f.data_inizio else None,
             "data_fine_prevista": str(f.data_fine_prevista) if f.data_fine_prevista else None}
            for f in fasi_da_mostrare
        ],
        "note_condivise": [
            {"id": n.id, "data": str(n.data), "testo": n.attivita or "",
             "meteo": n.meteo or "", "fonte": n.fonte or "manuale"}
            for n in note
        ],
        "appuntamenti": [
            {"id": f.id, "nome": f.nome, "data": str(f.data_inizio),
             "data_fine": str(f.data_fine_prevista) if f.data_fine_prevista else None,
             "colore": f.colore, "stato": f.stato}
            for f in appuntamenti[:5]
        ],
    }

@router.post("/{cantiere_id}/fasi", response_model=FaseOut, status_code=201)
def crea_fase(cantiere_id: int, body: FaseCreate, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_staff(user)
    try:
        fase = FaseLavoro(cantiere_id=cantiere_id, **body.model_dump())
        db.add(fase); db.commit(); db.refresh(fase)
        return fase
    except Exception as e:
        db.rollback(); raise HTTPException(500, f"Errore: {e}")

@router.put("/{cantiere_id}/fasi/{fase_id}", response_model=FaseOut)
def aggiorna_fase(cantiere_id: int, fase_id: int, body: FaseUpdate, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_staff(user)
    fase = db.query(FaseLavoro).filter(FaseLavoro.id == fase_id, FaseLavoro.cantiere_id == cantiere_id).first()
    if not fase: raise HTTPException(404, "Non trovata")
    for k, v in body.model_dump(exclude_none=True).items(): setattr(fase, k, v)
    from datetime import date as d_today
    oggi = d_today.today()
    if fase.percentuale >= 100:
        fase.stato = "completata"
        if not fase.data_fine_reale: fase.data_fine_reale = oggi
    elif fase.data_fine_prevista and oggi > fase.data_fine_prevista and fase.percentuale < 100:
        fase.stato = "in_ritardo"
    elif fase.percentuale > 0:
        fase.stato = "in_corso"
    db.commit(); db.refresh(fase)
    # Ricalcola avanzamento del cantiere come media delle fasi
    tutte_fasi = db.query(FaseLavoro).filter(FaseLavoro.cantiere_id == cantiere_id).all()
    if tutte_fasi:
        cantiere_obj = db.query(Cantiere).filter(Cantiere.id == cantiere_id).first()
        if cantiere_obj:
            cantiere_obj.avanzamento = round(sum(f.percentuale for f in tutte_fasi) / len(tutte_fasi), 1)
            db.commit()
    return fase

@router.delete("/{cantiere_id}/fasi/{fase_id}", status_code=204)
def elimina_fase(cantiere_id: int, fase_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user); _solo_staff(user)
    fase = db.query(FaseLavoro).filter(FaseLavoro.id == fase_id, FaseLavoro.cantiere_id == cantiere_id).first()
    if not fase: raise HTTPException(404, "Non trovata")
    db.delete(fase); db.commit()


# ─── HELPER JSON REPAIR ───────────────────────────────────────────────────────

def _ripara_json_array(testo: str) -> str:
    """
    Se Claude ha troncato il JSON per limite token, chiude l'array
    all'ultimo oggetto completo valido.
    """
    testo = testo.strip()
    # Rimuove markdown se presente
    if testo.startswith("```"):
        parti = testo.split("```")
        testo = parti[1] if len(parti) > 1 else testo
        if testo.startswith("json"):
            testo = testo[4:]
        testo = testo.strip()

    if not testo.startswith("["):
        return testo  # non è un array, lascia invariato

    # Prova parse diretto prima
    import json as _json
    try:
        _json.loads(testo)
        return testo
    except Exception:
        pass

    # Taglia all'ultima `}` seguita da eventuale `,` e chiudi con `]`
    ultimo_oggetto = testo.rfind("}")
    if ultimo_oggetto == -1:
        return "[]"
    troncato = testo[:ultimo_oggetto + 1].rstrip().rstrip(",")
    riparato = troncato + "\n]"
    try:
        _json.loads(riparato)
        return riparato
    except Exception:
        return "[]"


# ─── IMPORT COMPUTO DA FILE (AI) ──────────────────────────────────────────────

@router.post("/{cantiere_id}/preventivi/import-computo")
async def import_computo_ai(
    cantiere_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: Utente = Depends(get_current_user),
):
    """
    Riceve Excel/CSV/PDF con computo metrico.
    Claude interpreta la struttura e restituisce le voci pronte per il preventivo.
    """
    from app.config import settings
    _check(cantiere_id, db, user)
    _solo_economia(user)

    if not settings.ANTHROPIC_API_KEY:
        raise HTTPException(503, "Anthropic API key non configurata")

    contenuto = await file.read()
    nome = (file.filename or "").lower()
    ext = os.path.splitext(nome)[1]

    import re as _re, json as _json

    def _parse_numero(val):
        """Converte '35,00 €', '1.234,56', 1234.56 → float oppure None"""
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip().replace('€','').replace(' ','').replace('\xa0','')
        if ',' in s and '.' in s:
            s = s.replace('.','').replace(',','.')
        elif ',' in s:
            s = s.replace(',','.')
        s = _re.sub(r'[^\d.\-]','',s)
        try:
            return float(s) if s else None
        except Exception:
            return None

    PAROLE_SKIP = {"totale","totale generale","subtotale","totale complessivo",
                   "importo complessivo","sommano","totale parziale","totale ivato",
                   "totale complessivo ivato"}
    PAROLE_INTESTAZIONE = {"descrizione","voce","art.","lavorazione","u.m.","unità"}
    KEYWORD_RIEPILOGO = ["importo totale","importo complessivo","iva ","iva\n",
                         "totale ivato","totale complessivo ivato","sommano euro"]
    # Keyword Python che marcano una riga sospetta (subtotali, alternativi, ecc.)
    KW_SOSPETTA = [
        "totale", "subtotale", "sommano", "importo", "riepilogo",
        "alternativ", "oppure", "in alternativa", "variante", "opzione",
        "cap.", "capitolo", "fase ", "sezione", "parte ",
        "a -", "b -", "c -", "d -",  # intestazioni tipo "A - OPERE STRUTTURALI"
    ]

    def _is_sospetta_py(desc: str) -> bool:
        d = desc.lower().strip()
        return any(k in d for k in KW_SOSPETTA)

    voci = []

    try:
        if ext in (".xlsx", ".xls"):
            import openpyxl, io as _io
            wb = openpyxl.load_workbook(_io.BytesIO(contenuto), data_only=True)
            ws = wb.active
            n_col = ws.max_column

            # Rileva colonne dalle intestazioni (prime 10 righe)
            col_map = {}
            header_row_idx = None
            for rnum, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
                for ci, cell in enumerate(row):
                    if cell is None: continue
                    s = str(cell).lower().strip()
                    if any(k in s for k in ["descri","lavora","voce","oggetto"]) and "desc" not in col_map:
                        col_map["desc"] = ci
                    if any(k in s for k in ["u.m.","unità","misura"]) and "um" not in col_map:
                        col_map["um"] = ci
                    if any(k in s for k in ["quant","qta","q.tà"]) and "qt" not in col_map:
                        col_map["qt"] = ci
                    if any(k in s for k in ["prezzo","p.u.","unitario","costo unit"]) and "prezzo" not in col_map:
                        col_map["prezzo"] = ci
                    if any(k in s for k in ["totale","importo"]) and "totale" not in col_map:
                        col_map["totale"] = ci
                    if any(k in s for k in ["ricarico","ricaric","markup"]) and "ricarico" not in col_map:
                        col_map["ricarico"] = ci
                    if any(k in s for k in ["categ"]) and "cat" not in col_map:
                        col_map["cat"] = ci
                if col_map:
                    header_row_idx = rnum
                    break

            # Fallback: struttura offerta STEELEX (col0=codice,1=desc,2=um,3=qt,4=prezzo,5=totale)
            if not col_map and n_col >= 5:
                col_map = {"codice":0,"desc":1,"um":2,"qt":3,"prezzo":4,"totale":5}

            # Raccogli descrizioni per categorizzazione batch con Claude
            desc_per_categoria = []

            for row in ws.iter_rows(values_only=True):
                if not any(c is not None for c in row):
                    continue

                def g(key):
                    i = col_map.get(key)
                    return row[i] if i is not None and i < len(row) else None

                codice   = g("codice")
                desc     = g("desc")
                um       = g("um")
                qt_raw   = g("qt")
                pr_raw   = g("prezzo")
                tot_raw  = g("totale")
                ric_raw  = g("ricarico")
                cat_raw  = g("cat")

                desc_str = str(desc or "").replace("\n"," ").strip()
                desc_low = desc_str.lower()

                # Salta righe completamente vuote
                if not desc_str and pr_raw is None and tot_raw is None:
                    continue
                if desc_low in PAROLE_INTESTAZIONE:
                    continue
                # Righe che sono chiaramente totali generali/IVA: skip totale
                if any(k in desc_low for k in KEYWORD_RIEPILOGO):
                    continue

                qt_n  = _parse_numero(qt_raw)
                pr_n  = _parse_numero(pr_raw)
                tot_n = _parse_numero(tot_raw)

                # Righe subtotale di capitolo (descrizione "TOTALE xxx" con un importo ma senza qt/prezzo unitario)
                gia_sospetta = False
                if desc_low in PAROLE_SKIP and not codice:
                    if tot_n is not None:
                        # Teniamola ma la marchiamo come sospetta (è un subtotale di capitolo)
                        gia_sospetta = True
                    else:
                        continue

                # Righe intestazione di capitolo senza prezzi → sospette ma utili per contesto
                if pr_n is None and tot_n is None:
                    if desc_str and len(desc_str) > 3:
                        gia_sospetta = True
                        # inserisci come riga sospetta senza numeri
                        voci.append({
                            "id": len(voci) + 1,
                            "descrizione": desc_str[:200],
                            "categoria": "Altro",
                            "um": "-",
                            "quantita": 0.0,
                            "prezzo_costo": 0.0,
                            "ricarico_perc": 0.0,
                            "prezzo_cliente": 0.0,
                            "totale_costo": 0.0,
                            "totale_cliente": 0.0,
                            "sospetta": True,
                        })
                        desc_per_categoria.append(desc_str[:150])
                    continue

                # Calcola totale: usa il valore della colonna se c'è, altrimenti calcola
                if tot_n is not None:
                    totale_val = round(tot_n, 2)
                elif qt_n is not None and pr_n is not None:
                    totale_val = round(qt_n * pr_n, 2)
                elif pr_n is not None:
                    totale_val = round(pr_n, 2)
                else:
                    totale_val = 0.0

                # Prezzo unitario: usa colonna, oppure calcola da totale/qt
                if pr_n is not None:
                    prezzo_val = round(pr_n, 2)
                elif qt_n and qt_n != 0 and tot_n is not None:
                    prezzo_val = round(tot_n / qt_n, 2)
                else:
                    prezzo_val = totale_val

                qt_val = round(qt_n, 4) if qt_n is not None else 1.0

                ric_val = round(_parse_numero(ric_raw) or 0.0, 2)
                cat_val = str(cat_raw or "").strip() or "Altro"
                prezzo_cliente_val = round(prezzo_val * (1 + ric_val/100), 2)
                totale_cliente_val = round(qt_val * prezzo_cliente_val, 2)

                voci.append({
                    "id": len(voci) + 1,
                    "descrizione": desc_str[:200] or f"Voce {len(voci)+1}",
                    "categoria": cat_val,
                    "um": str(um or "").strip() or "cad",
                    "quantita": qt_val,
                    "prezzo_costo": prezzo_val,
                    "ricarico_perc": ric_val,
                    "prezzo_cliente": prezzo_cliente_val,
                    "totale_costo": totale_val,
                    "totale_cliente": totale_cliente_val,
                    "sospetta": gia_sospetta or _is_sospetta_py(desc_str),
                })
                desc_per_categoria.append(desc_str[:150])

            # Claude: categorizzazione + rilevamento righe sospette in un solo passaggio
            if voci and settings.ANTHROPIC_API_KEY:
                import anthropic
                claude = anthropic.Anthropic(api_key=settings.ANTHROPIC_API_KEY)
                lista_desc = "\n".join(f"{i+1}. {d}" for i,d in enumerate(desc_per_categoria))
                prompt_analisi = f"""Sei un esperto di computi metrici edili.
Analizza le seguenti voci estratte da un Excel e per CIASCUNA restituisci categoria e se è sospetta.

Categorie: Materiali, Manodopera, Nolo, Servizi, Sicurezza, Struttura, Finiture, Impianti, Ponteggi, Altro

Una riga è SOSPETTA (sospetta:true) se:
- È un subtotale/totale parziale di capitolo (es. "Totale struttura", "Sommano €", "Subtotale")
- È un'intestazione di capitolo/sezione (es. "A - OPERE STRUTTURALI", "CAP. 1")
- Contiene parole come: alternativa, oppure, variante, opzione, in sostituzione
- È un doppione della voce precedente con descrizione quasi identica
- È una nota generica senza prezzo reale

Voci (una per riga, formato: numero. descrizione):
{lista_desc[:8000]}

Rispondi ESCLUSIVAMENTE con un array JSON valido, senza markdown, senza testo aggiuntivo.
Formato esatto: [{{"cat":"Struttura","sospetta":false}},{{"cat":"Materiali","sospetta":true}}]
L'array deve avere ESATTAMENTE {len(desc_per_categoria)} elementi nello stesso ordine delle voci."""
                try:
                    msg = claude.messages.create(
                        model="claude-haiku-4-5-20251001", max_tokens=4000,
                        messages=[{"role":"user","content":prompt_analisi}]
                    )
                    raw = msg.content[0].text.strip()
                    # Rimuovi eventuali blocchi markdown ```json ... ```
                    raw = _re.sub(r'^```[a-z]*\s*', '', raw, flags=_re.MULTILINE)
                    raw = _re.sub(r'```\s*$', '', raw, flags=_re.MULTILINE).strip()
                    # Estrai array JSON (anche se Claude aggiunge testo prima/dopo)
                    m = _re.search(r'\[.*\]', raw, _re.DOTALL)
                    if m:
                        analisi = _json.loads(m.group())
                        for i, a in enumerate(analisi):
                            if i < len(voci) and isinstance(a, dict):
                                if isinstance(a.get("cat"), str):
                                    voci[i]["categoria"] = a["cat"]
                                # OR: se Python l'ha già marcata sospetta, non la resetta
                                voci[i]["sospetta"] = voci[i]["sospetta"] or bool(a.get("sospetta", False))
                    else:
                        print(f"[import-computo] Claude non ha restituito JSON valido: {raw[:200]}")
                except Exception as _e_claude:
                    print(f"[import-computo] Claude error: {_e_claude}")

        elif ext == ".csv":
            import csv, io as _io
            reader = csv.reader(_io.StringIO(contenuto.decode("utf-8", errors="replace")))
            righe = list(reader)
            # Trova riga intestazione
            for i_row, r in enumerate(righe[:10]):
                joined = " ".join(r).lower()
                if any(k in joined for k in ["descri","prezzo","totale","quantit"]):
                    righe = righe[i_row+1:]
                    break
            for r in righe[:300]:
                if len(r) < 3: continue
                desc_str = r[0].strip()
                qt_n  = _parse_numero(r[2] if len(r)>2 else None)
                pr_n  = _parse_numero(r[3] if len(r)>3 else None)
                tot_n = _parse_numero(r[4] if len(r)>4 else None)
                if not desc_str or (pr_n is None and tot_n is None): continue
                tot_v = round(tot_n or (qt_n or 1)*( pr_n or 0), 2)
                pr_v  = round(pr_n or (tot_n/qt_n if qt_n else tot_v), 2)
                voci.append({
                    "id": len(voci)+1, "descrizione": desc_str[:200],
                    "categoria": "Altro", "um": r[1].strip() if len(r)>1 else "cad",
                    "sospetta": False,
                    "quantita": round(qt_n or 1, 4), "prezzo_costo": pr_v,
                    "ricarico_perc": 0.0, "prezzo_cliente": pr_v,
                    "totale_costo": tot_v, "totale_cliente": tot_v,
                })

        elif ext == ".pdf":
            # PDF: renderizza pagine come immagini → Claude Vision legge le tabelle
            if not settings.ANTHROPIC_API_KEY:
                raise HTTPException(503, "Anthropic API key non configurata")

            import fitz as _fitz, base64 as _b64, anthropic as _ant, json as _json2

            doc_pdf = _fitz.open(stream=contenuto, filetype="pdf")

            # Prova prima estrazione testo; se c'è abbastanza testo usa testo, altrimenti vision
            testo_pagine = [p.get_text() for p in doc_pdf]
            testo_totale = "\n".join(testo_pagine)
            usa_vision = len(testo_totale.strip()) < 500  # meno di 500 char = PDF grafico

            claude_pdf = _ant.Anthropic(api_key=settings.ANTHROPIC_API_KEY)

            # Prompt compatto: usa valori brevi per risparmiare token
            PROMPT_VOCI = """Sei un esperto di computi metrici edili. Estrai TUTTE le voci di lavoro/fornitura con prezzo.

Formato risposta — array JSON, ogni oggetto:
{"d":"descrizione max 150 char","u":"um","q":quantita,"p":prezzo_unitario,"t":totale,"c":"categoria","s":bool_sospetta}

Categorie: Materiali Manodopera Nolo Servizi Sicurezza Struttura Finiture Impianti Ponteggi Altro
s=true se: subtotale, intestazione capitolo, voce alternativa, doppione.
IGNORA: copertine, foto, testi descrittivi, condizioni generali, firme.

Rispondi SOLO con l'array JSON, senza markdown, senza testo aggiuntivo."""

            def _parse_voci_pdf(raw):
                """Estrae voci dal JSON di Claude, tollerante a troncamenti."""
                raw = _re.sub(r'^```[a-z]*\s*|\s*```$', '', raw.strip(), flags=_re.MULTILINE).strip()
                # Prova parsing completo
                try:
                    m = _re.search(r'\[.*\]', raw, _re.DOTALL)
                    if m:
                        return _json2.loads(m.group())
                except Exception:
                    pass
                # JSON troncato: estrai oggetti completi uno per uno
                risultati = []
                for mo in _re.finditer(r'\{[^{}]+\}', raw, _re.DOTALL):
                    try:
                        risultati.append(_json2.loads(mo.group()))
                    except Exception:
                        pass
                return risultati

            try:
                if usa_vision:
                    mat = _fitz.Matrix(1.5, 1.5)
                    content_blocks = [{"type": "text", "text": PROMPT_VOCI + "\n\nPagine del documento:"}]
                    for pg_idx, page in enumerate(doc_pdf):
                        if pg_idx >= 20:
                            break
                        pix = page.get_pixmap(matrix=mat)
                        img_b64 = _b64.b64encode(pix.tobytes("png")).decode()
                        content_blocks.append({
                            "type": "image",
                            "source": {"type": "base64", "media_type": "image/png", "data": img_b64}
                        })
                    model_pdf = "claude-sonnet-4-6"
                else:
                    content_blocks = [{"type": "text", "text": PROMPT_VOCI + f"\n\nTesto documento:\n{testo_totale[:14000]}"}]
                    model_pdf = "claude-haiku-4-5-20251001"

                doc_pdf.close()

                msg_pdf = claude_pdf.messages.create(
                    model=model_pdf, max_tokens=16000,  # alto: evita troncamento
                    messages=[{"role": "user", "content": content_blocks}]
                )
                raw_pdf = msg_pdf.content[0].text.strip()
                parsed = _parse_voci_pdf(raw_pdf)
                if not parsed:
                    raise HTTPException(400, f"Claude non ha trovato voci nel documento. Risposta: {raw_pdf[:400]}")
                for v in parsed:
                    if not isinstance(v, dict):
                        continue
                    # Supporta sia chiavi lunghe ("descrizione") che corte ("d")
                    desc = str(v.get("d") or v.get("descrizione") or "").strip()
                    if not desc:
                        continue
                    qt  = float(v.get("q") or v.get("quantita") or 1.0)
                    pr  = float(v.get("p") or v.get("prezzo_cliente") or 0.0)
                    tot = float(v.get("t") or v.get("totale_cliente") or 0.0)
                    if tot == 0 and pr > 0:
                        tot = round(qt * pr, 2)
                    voci.append({
                        "id": len(voci) + 1,
                        "descrizione":    desc[:200],
                        "categoria":      str(v.get("c") or v.get("categoria") or "Altro"),
                        "um":             str(v.get("u") or v.get("um") or "cad"),
                        "quantita":       round(qt, 4),
                        "prezzo_costo":   round(pr, 2),
                        "ricarico_perc":  0.0,
                        "prezzo_cliente": round(pr, 2),
                        "totale_costo":   round(tot, 2),
                        "totale_cliente": round(tot, 2),
                        "sospetta":       bool(v.get("s") or v.get("sospetta", False)),
                    })
            except HTTPException:
                raise
            except Exception as e:
                raise HTTPException(400, f"Errore analisi PDF: {e}")

        else:
            raise HTTPException(400, "Formato non supportato. Carica un file .xlsx, .xls, .csv o .pdf")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Impossibile leggere il file: {e}")

    if not voci:
        raise HTTPException(400, "Nessuna voce trovata nel file. Controlla il formato.")

    return {"voci": voci, "totale_voci": len(voci)}


# ─── TEMPLATE EXCEL COMPUTO ───────────────────────────────────────────────────

@router.get("/template-computo")
def scarica_template_computo():
    """Scarica il template Excel standard per l'import del computo"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io as _io

    wb = openpyxl.Workbook()

    # ── Foglio istruzioni (primo, così si apre subito)
    wi = wb.active
    wi.title = "LEGGI PRIMA"
    wi.column_dimensions["A"].width = 80

    arancio = PatternFill("solid", fgColor="FF6B00")
    giallo  = PatternFill("solid", fgColor="FFF9C4")
    verde   = PatternFill("solid", fgColor="E8F5E9")

    istruzioni = [
        ("TEMPLATE IMPORT COMPUTO — STEELEX Cantieri", True, arancio),
        ("", False, None),
        ("REGOLA FONDAMENTALE:", True, giallo),
        ("Una riga = una voce. Niente sub-totali, niente totali generali, niente alternative.", False, giallo),
        ("Il sistema calcola i totali da solo.", False, giallo),
        ("", False, None),
        ("COLONNE OBBLIGATORIE:", True, None),
        ("  • Descrizione voce  →  testo libero", False, None),
        ("  • Quantità           →  numero (es: 120 oppure 1)", False, None),
        ("  • U.M.               →  mq / ml / mc / cad / h / kg / corpo / mese", False, None),
        ("  • Prezzo unitario €  →  prezzo di COSTO (quello che paghi tu)", False, None),
        ("", False, None),
        ("COLONNE FACOLTATIVE:", True, None),
        ("  • Categoria          →  Materiali / Manodopera / Nolo / Servizi / Sicurezza / Struttura / Finiture / Impianti / Ponteggi / Altro", False, None),
        ("  • Ricarico %         →  es. 30 (lascia vuoto per usare il ricarico globale del preventivo)", False, None),
        ("", False, None),
        ("COSA NON METTERE:", True, giallo),
        ("  ✗  Righe 'Totale', 'Subtotale', 'Importo totale', 'IVA'", False, giallo),
        ("  ✗  Righe vuote tra le voci (vanno ignorate ma confondono)", False, giallo),
        ("  ✗  Voci duplicate o alternative (inserisci solo quelle che vuoi importare)", False, giallo),
        ("", False, None),
        ("COME USARLO:", True, verde),
        ("  1. Vai nel foglio 'Computo'", False, verde),
        ("  2. Cancella le righe di esempio (righe 2-6)", False, verde),
        ("  3. Inserisci le tue voci — una per riga", False, verde),
        ("  4. Salva il file", False, verde),
        ("  5. Nell'app: sezione Economia → Computo → 'Importa AI' → seleziona questo file", False, verde),
    ]
    for rnum, (testo, bold, fill) in enumerate(istruzioni, 1):
        c = wi.cell(row=rnum, column=1, value=testo)
        c.font = Font(bold=bold, size=11 if bold else 10,
                      color="FFFFFF" if fill == arancio else "000000")
        if fill:
            c.fill = fill

    # ── Foglio computo
    ws = wb.create_sheet("Computo")

    # Intestazioni — solo le 6 colonne essenziali
    headers = [
        ("Descrizione voce", 50),
        ("Categoria", 16),
        ("U.M.", 8),
        ("Quantità", 10),
        ("Prezzo unitario €", 18),
        ("Ricarico %", 12),
    ]
    hfill = PatternFill("solid", fgColor="FF6B00")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    thin  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    ws.row_dimensions[1].height = 22
    for ci, (h, w) in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin
        ws.column_dimensions[chr(64+ci)].width = w

    # Righe di esempio (cancellabili)
    efill = PatternFill("solid", fgColor="FFF3E0")
    esempi = [
        ["Allestimento cantiere e segnaletica di sicurezza", "Sicurezza", "corpo", 1, 2000.00, 0],
        ["Ponteggio di facciata lato strada", "Ponteggi", "mq", 292.9, 27.00, 25],
        ["Rimozione manto di copertura in coppi di laterizio", "Struttura", "mq", 325.2, 11.00, 20],
        ["Nuova copertura in legno (orditura primaria e secondaria + tavolato)", "Struttura", "mq", 170.8, 140.00, 30],
        ["Membrana impermeabilizzante bituminosa sp. 4mm", "Materiali", "mq", 357.72, 14.80, 25],
    ]
    for ri, riga in enumerate(esempi, 2):
        for ci, val in enumerate(riga, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = efill; c.border = thin
            if ci >= 4:
                c.alignment = Alignment(horizontal="right")

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_computo_steelex.xlsx"}
    )


# ─── IMPORT FATTURA/BOLLA DA FOTO (AI VISION) ─────────────────────────────────

@router.post("/{cantiere_id}/spese/import-foto")
async def import_spesa_da_foto(
    cantiere_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: Utente = Depends(get_current_user),
):
    """
    Riceve foto o screenshot di fattura/bolla.
    Claude Vision estrae i dati e pre-compila il form spesa.
    """
    import base64
    from app.config import settings
    _check(cantiere_id, db, user)
    _solo_economia(user)

    if not settings.ANTHROPIC_API_KEY:
        raise HTTPException(503, "Anthropic API key non configurata")

    contenuto = await file.read()
    nome = (file.filename or "").lower()
    ext = os.path.splitext(nome)[1].lower()

    # Converti PDF prima pagina in immagine se necessario
    if ext == ".pdf":
        try:
            import fitz
            doc = fitz.open(stream=contenuto, filetype="pdf")
            pix = doc[0].get_pixmap(dpi=150)
            contenuto = pix.tobytes("png")
            ext = ".png"
        except Exception as e:
            raise HTTPException(400, f"Impossibile leggere PDF: {e}")

    media_type_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".webp": "image/webp"}
    media_type = media_type_map.get(ext, "image/jpeg")
    img_b64 = base64.standard_b64encode(contenuto).decode()

    import anthropic
    claude = anthropic.Anthropic(api_key=settings.ANTHROPIC_API_KEY)

    prompt = """Sei un assistente per la gestione amministrativa di cantieri edili italiani.
Analizza questa immagine di una fattura, bolla di consegna, ricevuta o documento contabile.

Estrai le seguenti informazioni e rispondile SOLO come JSON (senza markdown, senza testo aggiuntivo):

{
  "tipo_documento": "fattura" | "bolla" | "ricevuta" | "altro",
  "numero_documento": "numero fattura/bolla se presente",
  "data": "YYYY-MM-DD (se presente, altrimenti null)",
  "fornitore": "nome del fornitore/venditore",
  "descrizione": "descrizione sintetica di cosa è stato acquistato/fornito (max 100 caratteri)",
  "categoria": una tra ["materiali","manodopera","nolo","servizi","trasporto","altro"],
  "importo_netto": numero in euro (senza IVA, se distinguibile),
  "iva_perc": percentuale IVA (es: 22),
  "importo_totale": numero in euro (totale con IVA),
  "note": "eventuali informazioni utili aggiuntive (numero ordine, riferimenti, ecc.)"
}

Se un campo non è leggibile o non presente, usa null.
Per importo_totale usa il totale finale del documento.
Rispondi SOLO con il JSON, nulla altro."""

    try:
        msg = claude.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1024,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_b64}},
                    {"type": "text", "text": prompt},
                ]
            }]
        )
        testo = msg.content[0].text.strip()
        if testo.startswith("```"):
            testo = testo.split("```")[1]
            if testo.startswith("json"):
                testo = testo[4:]

        import json
        dati = json.loads(testo)

        # Normalizza importi
        for campo in ("importo_netto", "iva_perc", "importo_totale"):
            if dati.get(campo) is not None:
                try:
                    dati[campo] = round(float(str(dati[campo]).replace(",",".")), 2)
                except Exception:
                    dati[campo] = None

        return dati

    except Exception as e:
        raise HTTPException(500, f"Errore analisi Claude: {e}")


# ─── IMPORT GANTT DA FILE (AI) ────────────────────────────────────────────────

@router.post("/{cantiere_id}/fasi/import-gantt")
async def import_gantt_ai(
    cantiere_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: Utente = Depends(get_current_user),
):
    """
    Riceve Excel/CSV/PDF/immagine con cronoprogramma o Gantt.
    Claude interpreta le fasi e restituisce la lista pronta per l'import.
    """
    import base64
    from app.config import settings
    _check(cantiere_id, db, user)
    _solo_staff(user)

    if not settings.ANTHROPIC_API_KEY:
        raise HTTPException(503, "Anthropic API key non configurata")

    contenuto = await file.read()
    nome = (file.filename or "").lower()
    ext = os.path.splitext(nome)[1].lower()

    import anthropic
    claude = anthropic.Anthropic(api_key=settings.ANTHROPIC_API_KEY)

    PROMPT_BASE = """Sei un esperto di pianificazione cantieri edili italiani.
Analizza il cronoprogramma o diagramma di Gantt fornito ed estrai tutte le fasi di lavoro.

Per ogni fase restituisci un JSON con questi campi:
- nome: nome della fase (stringa)
- categoria: una tra [lavorazione, fornitura, collaudo, amministrativo, impianti, struttura, finiture]
- data_inizio: "YYYY-MM-DD" oppure null
- data_fine_prevista: "YYYY-MM-DD" oppure null
- percentuale: avanzamento 0-100 (default 0 se non indicato)
- stato: uno tra [pianificata, in_corso, completata, in_ritardo, sospesa] (default "pianificata")
- note: eventuali note aggiuntive (o null)
- ordine: numero progressivo che rispetta l'ordine originale (1, 2, 3...)

REGOLE:
- Rispetta l'ordine originale delle fasi
- Interpreta date in qualsiasi formato (gg/mm/aaaa, mm/aaaa, settimana N ecc.) e convertile in YYYY-MM-DD
- Se ci sono solo settimane (S1, S2...) o mesi relativi, calcola date assolute a partire da oggi
- Ignora righe di intestazione, totali, legende
- Rispondi SOLO con il JSON array, senza markdown, senza testo aggiuntivo"""

    try:
        # Immagine (foto Gantt, screenshot) → Vision
        if ext in (".jpg", ".jpeg", ".png", ".webp"):
            media_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".webp": "image/webp"}
            img_b64 = base64.standard_b64encode(contenuto).decode()
            msg = claude.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=8192,
                messages=[{"role": "user", "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_map[ext], "data": img_b64}},
                    {"type": "text", "text": PROMPT_BASE},
                ]}]
            )

        else:
            # File testuale: Excel, CSV, PDF
            testo_grezzo = ""
            if ext in (".xlsx", ".xls"):
                import openpyxl, io as _io
                wb = openpyxl.load_workbook(_io.BytesIO(contenuto), data_only=True)
                ws = wb.active
                righe = []
                for row in ws.iter_rows(values_only=True):
                    cella = [str(c).strip() if c is not None else "" for c in row]
                    if any(cella):
                        righe.append("\t".join(cella))
                testo_grezzo = "\n".join(righe[:300])
            elif ext == ".csv":
                import csv, io as _io
                reader = csv.reader(_io.StringIO(contenuto.decode("utf-8", errors="replace")))
                testo_grezzo = "\n".join(["\t".join(r) for r in reader][:300])
            elif ext == ".pdf":
                import fitz
                # Prova prima text, poi screenshot della prima pagina
                doc = fitz.open(stream=contenuto, filetype="pdf")
                testo_grezzo = "\n".join(doc[i].get_text() for i in range(min(len(doc), 5)))
                # Se il PDF sembra grafico (poco testo), usa Vision sulla prima pagina
                if len(testo_grezzo.strip()) < 100:
                    pix = doc[0].get_pixmap(dpi=150)
                    img_bytes = pix.tobytes("png")
                    img_b64 = base64.standard_b64encode(img_bytes).decode()
                    msg = claude.messages.create(
                        model="claude-haiku-4-5-20251001",
                        max_tokens=8192,
                        messages=[{"role": "user", "content": [
                            {"type": "image", "source": {"type": "base64", "media_type": "image/png", "data": img_b64}},
                            {"type": "text", "text": PROMPT_BASE},
                        ]}]
                    )
                    testo_grezzo = None  # già usato vision
            else:
                testo_grezzo = contenuto.decode("utf-8", errors="replace")[:8000]

            if testo_grezzo is not None:
                msg = claude.messages.create(
                    model="claude-haiku-4-5-20251001",
                    max_tokens=8192,
                    messages=[{"role": "user", "content": f"{PROMPT_BASE}\n\nContenuto file:\n{testo_grezzo[:5000]}\n\nRisposta (solo JSON array):"}]
                )

        testo = msg.content[0].text.strip()
        if testo.startswith("```"):
            testo = testo.split("```")[1]
            if testo.startswith("json"):
                testo = testo[4:]

        import json
        testo = _ripara_json_array(testo)
        fasi = json.loads(testo)
        if not isinstance(fasi, list):
            raise ValueError("Non è una lista")

        COLORI = ['#FF6B00','#3b82f6','#22c55e','#ef4444','#8b5cf6','#f59e0b','#06b6d4','#ec4899']
        for i, f in enumerate(fasi):
            f["ordine"] = i + 1
            f.setdefault("colore", COLORI[i % len(COLORI)])
            f.setdefault("percentuale", 0)
            f.setdefault("stato", "pianificata")
            f.setdefault("nota", None)
            # Normalizza percentuale
            try:
                f["percentuale"] = min(100, max(0, int(float(str(f["percentuale"]).replace("%","")))))
            except Exception:
                f["percentuale"] = 0

        return {"fasi": fasi, "totale_fasi": len(fasi)}

    except Exception as e:
        raise HTTPException(500, f"Errore interpretazione Claude: {e}")


# ─── EXPORT EXCEL ─────────────────────────────────────────────────────────────

@router.get("/{cantiere_id}/export/excel")
def export_excel(cantiere_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    """Genera un file Excel con riepilogo economico, spese e SAL del cantiere."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    cantiere = _check(cantiere_id, db, user)
    _solo_economia(user)

    spese = db.query(Spesa).filter(Spesa.cantiere_id == cantiere_id).order_by(Spesa.data).all()
    sal_list = db.query(SAL).filter(SAL.cantiere_id == cantiere_id).order_by(SAL.numero).all()
    preventivi = db.query(PreventivoCantiere).filter(PreventivoCantiere.cantiere_id == cantiere_id).all()
    prev_ok = next((p for p in preventivi if p.stato == "accettato"), preventivi[0] if preventivi else None)

    wb = Workbook()

    # Stili
    arancio = "FF6B00"
    blu = "1A1A2E"
    header_font = Font(color="FFFFFF", bold=True, size=11)
    arancio_fill = PatternFill("solid", fgColor=arancio)
    blu_fill = PatternFill("solid", fgColor=blu)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    def _header_row(ws, cols, fill):
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = fill
            cell.alignment = center
            cell.border = border

    def _col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Foglio 1: Riepilogo ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Riepilogo"
    totale_speso = sum(s.importo for s in spese)
    sal_emessi = sum(s.importo for s in sal_list if s.stato in ("emesso", "pagato"))
    sal_pagati = sum(s.importo for s in sal_list if s.stato == "pagato")
    budget = prev_ok.subtotale if prev_ok else (cantiere.budget or 0)

    ws1.append(["STEELEX — Riepilogo Economico"])
    ws1["A1"].font = Font(bold=True, size=14, color=arancio)
    ws1.append([f"Cantiere: {cantiere.nome}"])
    ws1.append([f"Esportato il: {date.today().strftime('%d/%m/%Y')}"])
    ws1.append([])

    righe = [
        ("Budget preventivo (imponibile)", f"€ {budget:,.2f}"),
        ("Totale spese registrate", f"€ {totale_speso:,.2f}"),
        ("Margine atteso", f"€ {(budget - totale_speso):,.2f}"),
        ("SAL emessi", f"€ {sal_emessi:,.2f}"),
        ("SAL pagati (incassati)", f"€ {sal_pagati:,.2f}"),
        ("Da incassare", f"€ {(sal_emessi - sal_pagati):,.2f}"),
    ]
    for etichetta, valore in righe:
        ws1.append([etichetta, valore])
        ws1[ws1.max_row][0].font = Font(bold=True)

    _col_widths(ws1, [35, 22])

    # ── Foglio 2: Spese ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Spese")
    _header_row(ws2, ["#", "Data", "Descrizione", "Fornitore", "Categoria", "Importo €", "Note"], arancio_fill)
    for i, s in enumerate(spese, 1):
        ws2.append([
            i,
            s.data.strftime("%d/%m/%Y") if s.data else "",
            s.descrizione,
            s.fornitore or "",
            s.categoria,
            round(s.importo, 2),
            s.note or "",
        ])
        for cell in ws2[ws2.max_row]:
            cell.border = border
    # Totale
    ws2.append(["", "", "", "", "TOTALE", round(totale_speso, 2), ""])
    for cell in ws2[ws2.max_row]:
        cell.font = Font(bold=True)
        cell.border = border
    _col_widths(ws2, [5, 13, 35, 25, 15, 14, 30])

    # ── Foglio 3: SAL ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("SAL")
    _header_row(ws3, ["N°", "Titolo", "Data", "% SAL", "Importo €", "Stato", "Note"], blu_fill)
    for s in sal_list:
        ws3.append([
            s.numero,
            s.titolo,
            s.data.strftime("%d/%m/%Y") if s.data else "",
            f"{s.percentuale:.1f}%",
            round(s.importo, 2),
            s.stato.upper(),
            s.note or "",
        ])
        for cell in ws3[ws3.max_row]:
            cell.border = border
    _col_widths(ws3, [5, 35, 13, 10, 14, 12, 30])

    # ── Foglio 4: Computo preventivo ─────────────────────────────────────────
    if prev_ok and prev_ok.voci:
        ws4 = wb.create_sheet("Computo")
        _header_row(ws4, ["#", "Descrizione", "Categoria", "Um", "Qty", "P.Costo €", "Ricarico%", "P.Cliente €", "Tot.Costo €", "Tot.Cliente €"], arancio_fill)
        for i, v in enumerate(prev_ok.voci, 1):
            ws4.append([
                i,
                v.get("descrizione", ""),
                v.get("categoria", ""),
                v.get("um", ""),
                v.get("quantita", 0),
                round(v.get("prezzo_costo", 0), 2),
                f"{v.get('ricarico_perc', 0):.1f}%",
                round(v.get("prezzo_cliente", 0), 2),
                round(v.get("totale_costo", 0), 2),
                round(v.get("totale_cliente", 0), 2),
            ])
            for cell in ws4[ws4.max_row]:
                cell.border = border
        # Totali
        ws4.append(["", "", "", "", "", "", "", "SUBTOTALE", "", round(prev_ok.subtotale, 2)])
        ws4.append(["", "", "", "", "", "", "", f"IVA {prev_ok.iva_perc:.0f}%", "", round(prev_ok.totale - prev_ok.subtotale, 2)])
        ws4.append(["", "", "", "", "", "", "", "TOTALE IVA INCL.", "", round(prev_ok.totale, 2)])
        for row in ws4.iter_rows(min_row=ws4.max_row - 2, max_row=ws4.max_row):
            for cell in row:
                cell.font = Font(bold=True)
        _col_widths(ws4, [5, 35, 18, 7, 8, 13, 11, 13, 13, 14])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome_file = f"economico_{cantiere.nome.replace(' ', '_')}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_file}"'},
    )


# ─── PDF PREVENTIVO ───────────────────────────────────────────────────────────

@router.get("/{cantiere_id}/preventivi/{prev_id}/genera-pdf")
def genera_pdf_preventivo(cantiere_id: int, prev_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    """Genera PDF preventivo formattato STEELEX pronto per il cliente."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT

    cantiere = _check(cantiere_id, db, user)
    _solo_economia(user)
    prev = db.query(PreventivoCantiere).filter(
        PreventivoCantiere.id == prev_id,
        PreventivoCantiere.cantiere_id == cantiere_id
    ).first()
    if not prev:
        raise HTTPException(404, "Preventivo non trovato")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)

    ARANCIO = colors.HexColor("#FF6B00")
    BLU = colors.HexColor("#1A1A2E")
    GRIGIO = colors.HexColor("#F5F5F5")

    styles = getSampleStyleSheet()
    style_titolo = ParagraphStyle("titolo", parent=styles["Heading1"], textColor=ARANCIO, fontSize=22, spaceAfter=2)
    style_sub = ParagraphStyle("sub", parent=styles["Normal"], textColor=BLU, fontSize=10)
    style_label = ParagraphStyle("label", parent=styles["Normal"], textColor=BLU, fontSize=9, fontName="Helvetica-Bold")
    style_small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, textColor=colors.grey)
    style_right = ParagraphStyle("right", parent=styles["Normal"], alignment=TA_RIGHT, fontSize=9)
    style_note = ParagraphStyle("note", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    story = []

    # Intestazione
    story.append(Paragraph("STEELEX", style_titolo))
    story.append(Paragraph("Costruzioni Light Steel Frame", style_sub))
    story.append(HRFlowable(width="100%", thickness=2, color=ARANCIO, spaceAfter=6))

    # Info preventivo
    numero = prev.numero or f"PRV-{prev.id:04d}"
    data_str = prev.data.strftime("%d/%m/%Y") if prev.data else date.today().strftime("%d/%m/%Y")
    info_data = [
        [Paragraph(f"<b>PREVENTIVO N°</b> {numero}", style_label),
         Paragraph(f"<b>Data:</b> {data_str}", style_right)],
        [Paragraph(f"<b>Cantiere:</b> {cantiere.nome}", style_label),
         Paragraph(f"<b>Validità:</b> {prev.validita_giorni} giorni", style_right)],
    ]
    info_table = Table(info_data, colWidths=["60%", "40%"])
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), GRIGIO),
        ("ROWPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 6*mm))

    # Tabella voci
    story.append(Paragraph("COMPUTO METRICO ESTIMATIVO", style_label))
    story.append(Spacer(1, 2*mm))

    headers = ["#", "Descrizione", "Cat.", "Um", "Qty", "P.Unit. €", "Totale €"]
    col_widths = [10*mm, 65*mm, 22*mm, 12*mm, 14*mm, 22*mm, 22*mm]
    table_data = [headers]
    voci = prev.voci or []
    for i, v in enumerate(voci, 1):
        table_data.append([
            str(i),
            v.get("descrizione", ""),
            v.get("categoria", ""),
            v.get("um", "cad"),
            str(v.get("quantita", 1)),
            f"€ {v.get('prezzo_cliente', 0):,.2f}",
            f"€ {v.get('totale_cliente', 0):,.2f}",
        ])

    # Righe totali
    table_data.append(["", "", "", "", "", "Imponibile", f"€ {prev.subtotale:,.2f}"])
    table_data.append(["", "", "", "", "", f"IVA {prev.iva_perc:.0f}%", f"€ {(prev.totale - prev.subtotale):,.2f}"])
    table_data.append(["", "", "", "", "", "TOTALE", f"€ {prev.totale:,.2f}"])
    table_data.append(["", "", "", "", "", f"Acconto {prev.acconto_perc:.0f}%", f"€ {prev.acconto_importo:,.2f}"])

    n_voci = len(voci)
    voci_table = Table(table_data, colWidths=col_widths)
    ts = TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), BLU),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        # Righe voci
        ("FONTSIZE", (0, 1), (-1, n_voci), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, n_voci), [colors.white, GRIGIO]),
        ("ALIGN", (4, 1), (-1, n_voci), "RIGHT"),
        # Righe totali
        ("FONTNAME", (0, n_voci + 1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, n_voci + 1), (-1, -1), 9),
        ("ALIGN", (5, n_voci + 1), (6, -1), "RIGHT"),
        ("BACKGROUND", (0, n_voci + 3), (-1, n_voci + 3), ARANCIO),
        ("TEXTCOLOR", (0, n_voci + 3), (-1, n_voci + 3), colors.white),
        # Griglia
        ("GRID", (0, 0), (-1, n_voci), 0.3, colors.lightgrey),
        ("LINEABOVE", (0, n_voci + 1), (-1, n_voci + 1), 1, ARANCIO),
        ("ROWPADDING", (0, 0), (-1, -1), 4),
    ])
    voci_table.setStyle(ts)
    story.append(voci_table)
    story.append(Spacer(1, 6*mm))

    # Note
    if prev.note:
        story.append(Paragraph("<b>Note:</b>", style_label))
        story.append(Paragraph(prev.note, style_note))
        story.append(Spacer(1, 4*mm))

    # Piè di pagina firma
    story.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey))
    story.append(Spacer(1, 4*mm))
    firma_data = [
        [Paragraph("Per accettazione:", style_label), Paragraph("STEELEX — Fontana Raffaele Srl", style_label)],
        [Paragraph("_" * 35, style_small), Paragraph("_" * 35, style_small)],
        [Paragraph("Timbro e firma cliente", style_small), Paragraph("Firma", style_small)],
    ]
    firma_table = Table(firma_data, colWidths=["50%", "50%"])
    story.append(firma_table)

    doc.build(story)
    buf.seek(0)

    nome_file = f"preventivo_{numero}_{cantiere.nome.replace(' ', '_')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nome_file}"'},
    )


# ─── IMPORT SPESE DA EXCEL ────────────────────────────────────────────────────

@router.post("/{cantiere_id}/spese/import-excel")
async def import_spese_excel(
    cantiere_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: Utente = Depends(get_current_user),
):
    """
    Legge un file Excel con colonne: data, descrizione, fornitore, categoria, importo, note.
    Restituisce le righe parsate per preview — il frontend le conferma prima del salvataggio.
    Colonne accettate in modo flessibile (case-insensitive, alias italiani/inglesi).
    """
    _check(cantiere_id, db, user)
    _solo_economia(user)

    contenuto = await file.read()
    nome_file = (file.filename or "").lower()
    ext = os.path.splitext(nome_file)[1].lower()

    # Rilevamento formato da magic bytes (più affidabile dell'estensione)
    # XLS (OLE2): D0 CF 11 E0 — XLSX (ZIP): 50 4B 03 04 — CSV: testo
    def _is_xls(data): return len(data) >= 8 and data[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'
    def _is_xlsx(data): return len(data) >= 4 and data[:4] == b'PK\x03\x04'

    righe_raw = []
    try:
        if _is_xls(contenuto):
            # Vecchio formato .xls (OLE2) — usa xlrd
            try:
                import xlrd
                wb = xlrd.open_workbook(file_contents=contenuto)
                ws = wb.sheet_by_index(0)
                righe_raw = [ws.row_values(i) for i in range(ws.nrows)]
            except ImportError:
                raise HTTPException(503, "Libreria xlrd non installata sul server. Usa il formato .xlsx o .csv")
        elif _is_xlsx(contenuto) or ext in (".xlsx", ".xlsm", ".ods"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contenuto), data_only=True)
            ws = wb.active
            righe_raw = [list(row) for row in ws.iter_rows(values_only=True)]
        else:
            # CSV / testo (fallback)
            import csv
            try:
                import chardet
                enc = chardet.detect(contenuto)["encoding"] or "utf-8"
            except ImportError:
                enc = "utf-8"
            testo = contenuto.decode(enc, errors="replace")
            try:
                dialect = csv.Sniffer().sniff(testo[:2048], delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            reader = csv.reader(testo.splitlines(), dialect)
            righe_raw = [list(row) for row in reader]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Impossibile leggere il file ({ext or 'formato sconosciuto'}): {e}")

    ALIAS_COLONNE = {
        "data": ["data", "date", "data spesa", "data doc", "data documento", "giorno", "dt", "del"],
        "descrizione": ["descrizione", "description", "oggetto", "voce", "causale", "desc", "fornitore/descrizione", "oggetto/descrizione", "fattura", "documento"],
        "fornitore": ["fornitore", "supplier", "fornitore/venditore", "venditore", "azienda", "ditta", "ragione sociale", "nominativo"],
        "categoria": ["categoria", "category", "tipo", "type", "tipologia"],
        "importo": ["importo", "amount", "totale", "importo totale", "€", "euro", "spesa",
                    "imponibile", "netto", "importo netto", "tot", "tot.", "totale fattura",
                    "importo ivato", "importo lordo", "lordo", "valore",
                    "costo totale", "costo complessivo"],
        "note": ["note", "notes", "annotazioni", "commento", "commenti", "riferimento", "rif", "rif."],
    }

    CATEGORIE_VALIDE = {"materiali", "manodopera", "nolo", "servizi", "trasporto", "altro"}
    ALIAS_CATEGORIE = {
        "materiale": "materiali", "material": "materiali", "materials": "materiali",
        "mat": "materiali", "manodopera": "manodopera", "mano": "manodopera",
        "labor": "manodopera", "lavoro": "manodopera", "nolo": "nolo", "rent": "nolo",
        "noleggio": "nolo", "servizi": "servizi", "services": "servizi",
        "servizio": "servizi", "trasporto": "trasporto", "transport": "trasporto",
        "logistica": "trasporto", "altro": "altro", "other": "altro",
        "varie": "altro", "vario": "altro",
    }

    # Righe da skippare (totali/subtotali/intestazioni ripetute)
    SKIP_DESC = {"totale", "totale generale", "subtotale", "totale parziale", "totale complessivo",
                 "grand total", "sub total", "subtotal", "totali", "riepilogo", "totale iva"}

    def _parsa_numero(raw):
        """Converte stringa in float gestendo formato italiano (1.234,56) e inglese (1234.56)."""
        s = str(raw).replace("€", "").replace(" ", "").replace("\xa0", "").replace("+", "").strip()
        if not s or s in ("-", "—", "n/a", "na"):
            return None
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        try:
            return round(float(s), 2)
        except Exception:
            return None

    # Individua riga intestazione (prima riga con almeno 2 alias riconosciuti, oppure 1 se è "importo"+"descrizione")
    header_row_idx = None
    col_map = {}
    for i, row in enumerate(righe_raw):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        row_lower = [str(v).strip().lower() if v is not None else "" for v in row]
        matched = {}
        for campo, alias_list in ALIAS_COLONNE.items():
            for j, cell_val in enumerate(row_lower):
                if cell_val in alias_list:
                    matched[campo] = j
                    break
        n_match = len(matched)
        if n_match >= 2 or ("importo" in matched and "descrizione" in matched):
            header_row_idx = i
            col_map = matched
            break
        # Fallback: anche solo importo o descrizione (ma almeno 1 match)
        if n_match >= 1 and i < 5:
            header_row_idx = i
            col_map = matched
            # continua a cercare una riga con più match

    if header_row_idx is None:
        # Ultimo tentativo: individua colonne numeriche automaticamente
        for i, row in enumerate(righe_raw[:10]):
            numeric_cols = [j for j, v in enumerate(row) if v is not None and _parsa_numero(v) is not None]
            if numeric_cols:
                col_map = {"importo": numeric_cols[-1]}  # ultima colonna numerica = probabilmente il totale
                header_row_idx = max(0, i - 1)
                break

    if header_row_idx is None:
        raise HTTPException(400, "Struttura file non riconosciuta. Il file deve avere colonne leggibili (Descrizione, Importo, ecc.)")

    import logging as _log
    _log.warning(f"[IMPORT-EXCEL] header_row={header_row_idx} col_map={col_map} totale_righe={len(righe_raw)}")
    for _dbg_i in range(header_row_idx, min(header_row_idx + 5, len(righe_raw))):
        _log.warning(f"[IMPORT-EXCEL] riga[{_dbg_i}] = {[str(v)[:30] for v in righe_raw[_dbg_i]]}")

    righe = []
    errori = []
    for i, row in enumerate(righe_raw[header_row_idx + 1:], header_row_idx + 2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        def val(campo):
            idx = col_map.get(campo)
            if idx is None:
                return None
            v = row[idx] if idx < len(row) else None
            return str(v).strip() if v is not None else None

        # Salta righe totale/subtotale (controlla anche le altre celle della riga, non solo la descrizione)
        desc_check = (val("descrizione") or "").strip().lower()
        if desc_check in SKIP_DESC:
            continue
        row_text_lower = {str(v).strip().lower() for v in row if v is not None and str(v).strip()}
        if row_text_lower & SKIP_DESC:
            continue

        # Importo — usa _parsa_numero che gestisce formato italiano
        imp_raw = val("importo")
        if not imp_raw:
            continue
        importo = _parsa_numero(imp_raw)
        if importo is None:
            errori.append(f"Riga {i}: importo '{imp_raw}' non valido")
            continue
        if importo <= 0:
            continue

        # Data
        data_raw = val("data")
        data_str = None
        if data_raw:
            from datetime import datetime as _dt
            for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
                try:
                    data_str = _dt.strptime(data_raw, fmt).date().isoformat()
                    break
                except Exception:
                    pass
            if data_str is None:
                # Prova timestamp Excel (numero float)
                try:
                    from openpyxl.utils.datetime import from_excel
                    data_str = from_excel(float(data_raw)).date().isoformat()
                except Exception:
                    pass

        # Categoria
        cat_raw = (val("categoria") or "altro").lower().strip()
        categoria = ALIAS_CATEGORIE.get(cat_raw, "altro") if cat_raw not in CATEGORIE_VALIDE else cat_raw

        descrizione = val("descrizione") or "Spesa senza descrizione"
        fornitore = val("fornitore") or ""
        note = val("note") or ""

        righe.append({
            "riga_excel": i,
            "data": data_str or date.today().isoformat(),
            "descrizione": descrizione,
            "fornitore": fornitore,
            "categoria": categoria,
            "importo": importo,
            "note": note,
            "ok": True,
        })

    return {
        "righe": righe,
        "totale": round(sum(r["importo"] for r in righe), 2),
        "errori": errori,
        # Debug info (utile se 0 righe trovate)
        "debug": {
            "header_row": header_row_idx,
            "col_map": col_map,
            "prime_righe": [
                [str(v)[:40] if v is not None else "" for v in righe_raw[j]]
                for j in range(min(header_row_idx + 1, len(righe_raw)), min(header_row_idx + 6, len(righe_raw)))
            ] if header_row_idx is not None else [],
            "totale_righe_file": len(righe_raw),
        } if len(righe) == 0 else None,
    }


@router.post("/{cantiere_id}/spese/import-excel/conferma")
async def conferma_import_spese_excel(
    cantiere_id: int,
    righe: List[Any],
    db: Session = Depends(get_db),
    user: Utente = Depends(get_current_user),
):
    """Salva in bulk le spese selezionate dalla preview Excel."""
    _check(cantiere_id, db, user)
    _solo_economia(user)

    salvate = 0
    for r in righe:
        try:
            data_val = date.fromisoformat(r["data"]) if r.get("data") else date.today()
            sp = Spesa(
                cantiere_id=cantiere_id,
                descrizione=r.get("descrizione", "")[:500],
                fornitore=r.get("fornitore", "")[:255] or None,
                categoria=r.get("categoria", "altro"),
                importo=float(r.get("importo", 0)),
                data=data_val,
                note=r.get("note", "")[:1000] or None,
                creato_da=user.id,
            )
            db.add(sp)
            salvate += 1
        except Exception:
            continue
    db.commit()
    return {"salvate": salvate}


# ─── ORDINI ACQUISTO ─────────────────────────────────────────────────────────

class OrdineBody(BaseModel):
    fornitore_nome: str
    descrizione: str
    categoria: Optional[str] = "materiali"
    importo: float
    iva_perc: Optional[float] = 22.0
    stato: Optional[str] = "bozza"
    data_ordine: Optional[date] = None
    data_consegna_prevista: Optional[date] = None
    note: Optional[str] = None

def _ordine_dict(o: OrdineAcquisto) -> dict:
    return {
        "id": o.id,
        "cantiere_id": o.cantiere_id,
        "fornitore_nome": o.fornitore_nome,
        "descrizione": o.descrizione,
        "categoria": o.categoria,
        "importo": o.importo,
        "iva_perc": o.iva_perc,
        "importo_totale": round(o.importo * (1 + (o.iva_perc or 22) / 100), 2),
        "stato": o.stato,
        "data_ordine": o.data_ordine.isoformat() if o.data_ordine else None,
        "data_consegna_prevista": o.data_consegna_prevista.isoformat() if o.data_consegna_prevista else None,
        "note": o.note,
        "n_bolle": len(o.bolle) if hasattr(o, "bolle") else 0,
    }

@router.get("/{cantiere_id}/ordini")
def lista_ordini(cantiere_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user)
    _solo_economia(user)
    rows = db.query(OrdineAcquisto).filter(OrdineAcquisto.cantiere_id == cantiere_id).order_by(OrdineAcquisto.data_ordine.desc()).all()
    return [_ordine_dict(r) for r in rows]

@router.post("/{cantiere_id}/ordini")
def crea_ordine(cantiere_id: int, body: OrdineBody, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user)
    _solo_economia(user)
    row = OrdineAcquisto(
        cantiere_id=cantiere_id,
        fornitore_nome=body.fornitore_nome,
        descrizione=body.descrizione,
        categoria=body.categoria or "materiali",
        importo=body.importo,
        iva_perc=body.iva_perc or 22,
        stato=body.stato or "bozza",
        data_ordine=body.data_ordine or date.today(),
        data_consegna_prevista=body.data_consegna_prevista,
        note=body.note,
    )
    db.add(row); db.commit(); db.refresh(row)
    return _ordine_dict(row)

@router.put("/{cantiere_id}/ordini/{ordine_id}")
def aggiorna_ordine(cantiere_id: int, ordine_id: int, body: OrdineBody, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user)
    _solo_economia(user)
    row = db.query(OrdineAcquisto).filter(OrdineAcquisto.id == ordine_id, OrdineAcquisto.cantiere_id == cantiere_id).first()
    if not row:
        raise HTTPException(404)
    for k, v in body.dict(exclude_none=True).items():
        setattr(row, k, v)
    db.commit(); db.refresh(row)
    return _ordine_dict(row)

@router.patch("/{cantiere_id}/ordini/{ordine_id}/stato")
def aggiorna_stato_ordine(cantiere_id: int, ordine_id: int, stato: str, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user)
    _solo_economia(user)
    row = db.query(OrdineAcquisto).filter(OrdineAcquisto.id == ordine_id, OrdineAcquisto.cantiere_id == cantiere_id).first()
    if not row:
        raise HTTPException(404)
    stati_validi = {"bozza", "inviato", "confermato", "evaso", "annullato"}
    if stato not in stati_validi:
        raise HTTPException(422, f"Stato non valido. Valori: {stati_validi}")
    row.stato = stato
    db.commit()
    return _ordine_dict(row)

@router.delete("/{cantiere_id}/ordini/{ordine_id}")
def elimina_ordine(cantiere_id: int, ordine_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user)
    _solo_economia(user)
    row = db.query(OrdineAcquisto).filter(OrdineAcquisto.id == ordine_id, OrdineAcquisto.cantiere_id == cantiere_id).first()
    if not row:
        raise HTTPException(404)
    db.delete(row); db.commit()
    return {"ok": True}


# ─── BOLLE CONSEGNA (DDT) ────────────────────────────────────────────────────

class BollaBody(BaseModel):
    fornitore_nome: str
    numero_bolla: Optional[str] = None
    data: Optional[date] = None
    importo_stimato: Optional[float] = None
    descrizione: Optional[str] = None
    ordine_id: Optional[int] = None
    note: Optional[str] = None

def _bolla_dict(b: BollaConsegna) -> dict:
    return {
        "id": b.id,
        "cantiere_id": b.cantiere_id,
        "fornitore_nome": b.fornitore_nome,
        "numero_bolla": b.numero_bolla,
        "data": b.data.isoformat() if b.data else None,
        "importo_stimato": b.importo_stimato,
        "descrizione": b.descrizione,
        "ordine_id": b.ordine_id,
        "stato": b.stato,
        "foto_url": b.foto_url,
        "note": getattr(b, "note", None),
    }

@router.get("/{cantiere_id}/bolle")
def lista_bolle(cantiere_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user)
    _solo_economia(user)
    rows = db.query(BollaConsegna).filter(BollaConsegna.cantiere_id == cantiere_id).order_by(BollaConsegna.data.desc()).all()
    return [_bolla_dict(r) for r in rows]

@router.post("/{cantiere_id}/bolle")
def crea_bolla(cantiere_id: int, body: BollaBody, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user)
    _solo_economia(user)
    row = BollaConsegna(
        cantiere_id=cantiere_id,
        fornitore_nome=body.fornitore_nome,
        numero_bolla=body.numero_bolla,
        data=body.data or date.today(),
        importo_stimato=body.importo_stimato,
        descrizione=body.descrizione,
        ordine_id=body.ordine_id,
        stato="aperta",
    )
    db.add(row); db.commit(); db.refresh(row)
    return _bolla_dict(row)

@router.put("/{cantiere_id}/bolle/{bolla_id}")
def aggiorna_bolla(cantiere_id: int, bolla_id: int, body: BollaBody, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user)
    _solo_economia(user)
    row = db.query(BollaConsegna).filter(BollaConsegna.id == bolla_id, BollaConsegna.cantiere_id == cantiere_id).first()
    if not row:
        raise HTTPException(404)
    for k, v in body.dict(exclude_none=True).items():
        if hasattr(row, k):
            setattr(row, k, v)
    db.commit(); db.refresh(row)
    return _bolla_dict(row)

@router.post("/{cantiere_id}/bolle/{bolla_id}/foto")
async def upload_foto_bolla(cantiere_id: int, bolla_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user)
    _solo_economia(user)
    row = db.query(BollaConsegna).filter(BollaConsegna.id == bolla_id, BollaConsegna.cantiere_id == cantiere_id).first()
    if not row:
        raise HTTPException(404)
    url = await salva_file(file, f"bolle/{cantiere_id}/{bolla_id}")
    row.foto_url = url
    db.commit()
    return {"foto_url": url}

@router.delete("/{cantiere_id}/bolle/{bolla_id}")
def elimina_bolla(cantiere_id: int, bolla_id: int, db: Session = Depends(get_db), user: Utente = Depends(get_current_user)):
    _check(cantiere_id, db, user)
    _solo_economia(user)
    row = db.query(BollaConsegna).filter(BollaConsegna.id == bolla_id, BollaConsegna.cantiere_id == cantiere_id).first()
    if not row:
        raise HTTPException(404)
    db.delete(row); db.commit()
    return {"ok": True}
